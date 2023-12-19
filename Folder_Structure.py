import os
import openpyxl
import time
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import Alignment
import pandas as pd

def spreadsheet(target):
    dir_sheet = openpyxl.Workbook()
    dir_data = dir_sheet.active
    r = 2
    dir_data.cell(row=1, column=1, value="files/folders")
    dir_data.cell(row=1, column=2, value="Modification Date")
    dir_data.cell(row=1, column=3, value="Modification Time")
    dir_data.cell(row=1, column=4, value="Creation Date")
    dir_data.cell(row=1, column=5, value="Creation Time")
    dir_data.cell(row=1, column=6, value="file_extension")

    for root, dir, files in os.walk(target):
        try:
            for folder in dir:
                #print(folder)
                path = os.path.join(root, folder)
                creation_timestamp = os.path.getctime(path)
                modification_timestamp = os.path.getmtime(path)
                c_date = time.strftime('%m/%d/%Y', time.localtime(creation_timestamp))
                c_time = time.strftime('%H:%M:%S', time.localtime(creation_timestamp))
                m_date = time.strftime('%m/%d/%Y', time.localtime(modification_timestamp))
                m_time = time.strftime('%H:%M:%S', time.localtime(modification_timestamp))
                file_name, file_extension = os.path.splitext(path)

                dir_data.cell(row=r, column=1,value=path)
                dir_data.cell(row=r, column=2, value=m_date)
                dir_data.cell(row=r, column=3, value=m_time)
                dir_data.cell(row=r, column=4, value=c_date)
                dir_data.cell(row=r, column=5, value=c_time)
                dir_data.cell(row=r, column=6, value=file_extension)

                dir_data.cell(r, 2).alignment = Alignment(horizontal="center")
                dir_data.cell(r, 3).alignment = Alignment(horizontal="center")
                dir_data.cell(r, 4).alignment = Alignment(horizontal="center")
                dir_data.cell(r, 5).alignment = Alignment(horizontal="center")
                dir_data.cell(r, 6).alignment = Alignment(horizontal="center")
                r+=1
            for file in files:
                #print("          ", file)
                path = os.path.join(root, file)
                creation_timestamp = os.path.getctime(path)
                modification_timestamp = os.path.getmtime(path)
                c_date = time.strftime('%m/%d/%Y', time.localtime(creation_timestamp))
                c_time = time.strftime('%H:%M:%S', time.localtime(creation_timestamp))
                m_date = time.strftime('%m/%d/%Y', time.localtime(modification_timestamp))
                m_time = time.strftime('%H:%M:%S', time.localtime(modification_timestamp))
                file_name, file_extension = os.path.splitext(path)

                dir_data.cell(row=r, column=1, value=path)
                dir_data.cell(row=r, column=2, value=m_date)
                dir_data.cell(row=r, column=3, value=m_time)
                dir_data.cell(row=r, column=4, value=c_date)
                dir_data.cell(row=r, column=5, value=c_time)
                dir_data.cell(row=r, column=6, value=file_extension)


                r+=1

        except FileNotFoundError as fnf_error:
            pass
            r+=1

    dir_sheet.save(r"\\FTPCNTSK\EC_V_Drive\Anuj Singh\000-Proposed C6 Unit Folder Structure\Transmittal Received.xlsx")
    dir_sheet.close()

def beautify(target):

    path = os.path.join(target,"Transmittal Received.xlsx")
    dir_sheet = openpyxl.load_workbook(path)
    dir_data = dir_sheet.active
    dir_data.delete_cols(1)

    for col in range(2,dir_data.max_column+1):
        for row in range(2,dir_data.max_row+1):
            dir_data.cell(row, col).alignment = Alignment(horizontal="center")
            dir_data.cell(row, col).alignment = Alignment(horizontal="center")
            dir_data.cell(row, col).alignment = Alignment(horizontal="center")
            dir_data.cell(row, col).alignment = Alignment(horizontal="center")
            dir_data.cell(row, col).alignment = Alignment(horizontal="center")



    dir_data.cell(1, 2).alignment = Alignment(horizontal="center")
    dir_data.cell(1, 3).alignment = Alignment(horizontal="center")
    dir_data.cell(1, 4).alignment = Alignment(horizontal="center")
    dir_data.cell(1, 5).alignment = Alignment(horizontal="center")
    dir_data.cell(1, 6).alignment = Alignment(horizontal="center")

    dir_data.column_dimensions['A'].width = 50              #for width
    dir_data.column_dimensions['B'].width = 20
    dir_data.column_dimensions['C'].width = 20
    dir_data.column_dimensions['D'].width = 20
    dir_data.column_dimensions['E'].width = 20

    dir_data.freeze_panes = 'A2'                            #for freeze top row

    dir_data.cell(row = 1, column = 1).fill = PatternFill(start_color='00FFFF',end_color='00FFFF',fill_type='solid')    #for color on top
    dir_data.cell(row = 1, column = 2).fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
    dir_data.cell(row = 1, column = 3).fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
    dir_data.cell(row = 1, column = 4).fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
    dir_data.cell(row = 1, column = 5).fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
    dir_data.cell(row = 1, column = 6).fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')

    dir_sheet.save(path)
    dir_sheet.close()


def sortExcel():
    df = pd.read_excel(r"\\FTPCNTSK\EC_V_Drive\Anuj Singh\000-Proposed C6 Unit Folder Structure\Transmittal Received.xlsx")
    df_sorted = df.sort_values(by = ['Modification Date', "Modification Time"], ascending=False)
    df_sorted.to_excel(r"\\FTPCNTSK\EC_V_Drive\Anuj Singh\000-Proposed C6 Unit Folder Structure\Transmittal Received.xlsx")



target = r"\\FTPCNTSK\EC_V_Drive\Anuj Singh\000-Proposed C6 Unit Folder Structure\00-Transmittal"
spreadsheet(target)
sortExcel()
beautify(r"\\FTPCNTSK\EC_V_Drive\Anuj Singh\000-Proposed C6 Unit Folder Structure")
